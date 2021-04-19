import openpyxl

class Excel():

    def __init__(self, bk_name, sh_name):
        self.bk_name = bk_name
        self.sh_name = sh_name
        self.bk = openpyxl.load_workbook(self.bk_name)
        self.sh = self.bk['Sheet1']

    def getBook(self):
        return self.bk

    def getSheet(self):
        return self.sh

    def save(self):
        self.bk.save(self.bk_name)

    # 1行目の値が同じなら重複ありとする
    def isDuplicated(self, target_col):
        target_url = self.sh.cell(row=1,column=target_col).value

        for col in range(2,target_col):
            if self.sh.cell(row=1,column=col).value == target_url:
                return True

        return False

    # 2行目に値があれば取得済みとする
    def hasFetched(self, target_col):
        if self.sh.cell(row=2,column=target_col).value is None:
            return False
        else:
            return True
